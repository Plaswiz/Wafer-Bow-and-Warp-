import os
import re
import pandas as pd
import tkinter as tk
from tkinter import (
    filedialog, messagebox, Button, Label, Frame, Listbox,
    Scrollbar, Text, Entry, Checkbutton, IntVar, LabelFrame
)
from tkinter import ttk
from scipy.interpolate import griddata
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import plotly.graph_objects as go

from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from tkinter import font as tkfont  # For custom dialog fonts

# Configure logging
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")


# ------------------------------------------------------------------------------
# 1) "Smart" Extract Logic (left-to-right for '4')
# ------------------------------------------------------------------------------
def smart_extract_lot_wafer(path: str):
    logging.debug(f"[smart_extract_lot_wafer] Checking path: {path}")
    parts = re.split(r'[\\/]', path)
    logging.debug(f"Path parts: {parts}")

    if not parts:
        logging.debug("... no parts => (None,None)")
        return (None, None)

    # Define a regex pattern to match lot folders:
    # Starts with '4', followed by exactly six digits, and optionally followed by '_' and letters/numbers
    lot_folder_pattern = re.compile(r'^4\d{6}(?:[_]?\w*)?$', re.IGNORECASE)

    # Locate the lot using the refined pattern
    lot = None
    for part in parts:
        if lot_folder_pattern.match(part):
            lot = part
            logging.debug(f"Found lot subfolder matching pattern: {part}")
            break
    if not lot:
        logging.debug("... no matching lot folder => (None,None).")
        return (None, None)

    # Define a regex pattern to match any 'Run' folder (e.g., 'Run0001', 'Run0002', etc.)
    run_folder_pattern = re.compile(r'^Run\d+$', re.IGNORECASE)

    # Search for any 'RunXXXX' folder in the path
    wafer = None
    for i, part in enumerate(parts):
        if run_folder_pattern.match(part):
            if i > 0:
                wafer = parts[i - 1]
                logging.debug(f"Found wafer={wafer} before '{part}' folder")
                return (lot, wafer)
            else:
                logging.debug(f"'{part}' found at the start of path; no wafer folder.")
                return (lot, None)

    # If no Run folder found, return None
    logging.debug(f"No Run folder found => returning (lot={lot}, wafer={wafer})")
    return (lot, wafer)


# ------------------------------------------------------------------------------
# 2) "find_lot_wafer_upwards" – climb upward from file directory
# ------------------------------------------------------------------------------
def find_lot_wafer_upwards(start_dir: str):
    logging.info(f"[find_lot_wafer_upwards] Climb upward from: {start_dir}")
    current = start_dir

    run_folder_pattern = re.compile(r'^Run\d+$', re.IGNORECASE)  # Pattern to match 'Run' folders

    while True:
        if not current or not os.path.isdir(current):
            logging.debug(f"... {current} is invalid => (None,None)")
            return (None, None)

        basename = os.path.basename(current)
        logging.debug(f"Checking directory: {current}")

        # Updated lot folder pattern
        lot_folder_pattern = re.compile(r'^4\d{6}(?:[_]?\w*)?$', re.IGNORECASE)

        if lot_folder_pattern.match(basename):
            # Found lot folder
            lot = basename
            logging.debug(f"Found lot folder: {lot}")

            # Get relative path from lot folder to start_dir
            rel_path = os.path.relpath(start_dir, current)
            rel_parts = rel_path.split(os.sep)
            logging.debug(f"Relative path from lot to start_dir: {rel_parts}")

            # Search for any 'RunXXXX' folder in rel_parts
            wafer = None
            run_folder_pattern = re.compile(r'^Run\d+$', re.IGNORECASE)
            for i, part in enumerate(rel_parts):
                if run_folder_pattern.match(part):
                    if i > 0:
                        wafer = rel_parts[i - 1]
                        logging.debug(f"Found wafer={wafer} before '{part}' folder")
                        return (lot, wafer)
                    else:
                        logging.debug(f"'{part}' found at the start of relative path; no wafer folder.")
                        return (lot, None)
            # If no Run folder found in rel_path
            logging.debug(f"No 'RunXXXX' folder found in relative path from lot={lot} to start_dir={start_dir}")
            return (lot, None)

        # Move up one directory
        parent = os.path.dirname(current)
        if not parent or parent == current:
            logging.debug(f"Reached root from {current} => (None,None).")
            return (None, None)
        logging.debug(f"Going up to parent directory: {parent}")
        current = parent


# ------------------------------------------------------------------------------
# 3) Plane Fitting, Bow, Warp
# ------------------------------------------------------------------------------
def fit_plane(points, values, enable_radius=False, radius=95):
    """
    Fits a plane to the provided points and values.
    If enable_radius is True, uses only points outside the specified radius for fitting.
    """
    logging.debug("Starting plane fitting.")
    if enable_radius:
        # Select points outside the specified radius
        distances = np.linalg.norm(points, axis=1)
        mask = distances > radius
        outer_pts = points[mask]
        outer_vals = values[mask]
        logging.debug(f"Plane fitting with radius > {radius}: {len(outer_pts)} points selected.")
    else:
        # Use all points for plane fitting
        outer_pts = points
        outer_vals = values
        logging.debug("Plane fitting without radius filtering: using all points.")

    if len(outer_pts) < 3:
        logging.error("Not enough outer points to fit a plane.")
        raise ValueError("Insufficient outer points for plane fitting.")

    A = np.c_[outer_pts[:, 0], outer_pts[:, 1], np.ones(outer_pts.shape[0])]
    C, *_ = np.linalg.lstsq(A, outer_vals, rcond=None)
    logging.debug(f"Plane fitting coefficients: {C}")
    return C


def remove_plane(points, values, coeffs):
    """
    Removes the fitted plane from the values.
    """
    return values - (coeffs[0] * points[:, 0] + coeffs[1] * points[:, 1] + coeffs[2])


def calculate_bow_warp(points, corrected, enable_outlier_filter=False, compute_actual_center=False, z_threshold=3.0):
    """
    Calculates bow and warp from the corrected surface data.
    Applies outlier filtering to both bow and warp when enabled.
    """
    # Calculate centroid dynamically if enabled
    if compute_actual_center:
        centroid_x = np.median(points[:, 0])
        centroid_y = np.median(points[:, 1])
        logging.debug(f"Computed centroid: ({centroid_x}, {centroid_y})")
    else:
        centroid_x = 0
        centroid_y = 0
        logging.debug(f"Using default center: (0, 0)")

    # Interpolate corrected Z at centroid using 'linear' method
    center_val = griddata(points, corrected, (centroid_x, centroid_y), method='linear')

    if np.isnan(center_val):
        logging.warning(
            f"Interpolated center Z value is NaN for centroid ({centroid_x}, {centroid_y}). Setting bow to 0.")
        bow = 0.0
    else:
        logging.debug(f"Interpolated center Z value: {center_val}")
        # Calculate bow using mean to reduce outlier influence
        bow = (np.mean(corrected) - center_val) * 1000  # microns
        logging.debug(f"Calculated bow before filtering: {bow} microns")

    # Outlier Filtering for Bow and Warp Calculations
    if enable_outlier_filter:
        logging.debug("Outlier filtering enabled for bow and warp.")
        from scipy.stats import zscore
        z_scores = zscore(corrected)
        if np.isnan(z_scores).all():
            logging.warning("All Z-scores are NaN after filtering. Skipping outlier filtering.")
            filtered_corrected = corrected
            num_filtered = 0
        else:
            # Retain points within the z_threshold
            mask = np.abs(z_scores) < z_threshold
            filtered_corrected = corrected[mask]
            num_filtered = len(corrected) - len(filtered_corrected)
            logging.debug(f"Outlier filtering applied. Points excluded: {num_filtered}")

        if len(filtered_corrected) == 0:
            logging.warning("No points left after outlier filtering. Using unfiltered bow and warp values.")
            bow = (np.mean(corrected) - center_val) * 1000  # Recalculate without filtering
            warp = np.ptp(corrected) * 1000  # Recalculate without filtering
        else:
            bow = (np.mean(filtered_corrected) - center_val) * 1000  # Recalculate with filtered data
            logging.debug(f"Calculated bow after filtering: {bow} microns")
            warp = np.ptp(filtered_corrected) * 1000  # microns
            logging.debug(f"Calculated warp after filtering: {warp} microns")
    else:
        logging.debug("Outlier filtering not enabled.")
        # Calculate warp using all corrected points
        warp = np.ptp(corrected) * 1000  # microns
        logging.debug(f"Calculated warp: {warp} microns")

    return bow, warp


# ------------------------------------------------------------------------------
# 4) Generate Heatmap
# ------------------------------------------------------------------------------
def generate_heatmap(points, corrected_values_mm, bow, warp, lot, wafer, output_folder, interp_method):
    """
    Generates and saves a 2D heatmap for the wafer data.
    """
    corrected_um = corrected_values_mm * 1000.0
    grid_res = 200
    x = np.linspace(points[:, 0].min(), points[:, 0].max(), grid_res)
    y = np.linspace(points[:, 1].min(), points[:, 1].max(), grid_res)
    xg, yg = np.meshgrid(x, y)
    zg = griddata(points, corrected_um, (xg, yg), method=interp_method)

    fig = go.Figure(
        data=go.Heatmap(
            x=x,
            y=y,
            z=zg,
            colorscale='Viridis',
            colorbar=dict(title="Corrected Z [microns]", tickformat=".2f"),
        )
    )
    fig.update_layout(
        title=f"Wafer Heatmap - Lot: {lot}, Wafer: {wafer}<br>Bow: {bow:.1f} µm | Warp: {warp:.1f} µm",
        xaxis_title="X Coordinate",
        yaxis_title="Y Coordinate",
        xaxis=dict(scaleanchor="y", scaleratio=1),
        yaxis=dict(scaleanchor="x", scaleratio=1)
    )

    filename = f"{lot}_{wafer}_heatmap.html"
    out_file = os.path.join(output_folder, filename)
    fig.write_html(out_file)
    logging.info(f"Heatmap saved as {out_file}")


# ------------------------------------------------------------------------------
# 5) File Reading, Data Analysis
# ------------------------------------------------------------------------------
def sieve(topfolder, pattern=r'Run\d+\.txt'):
    """
    Recursively searches for files matching the given pattern within the topfolder.
    """
    matches = []
    for base, _, files in os.walk(topfolder):
        for f in files:
            if re.search(pattern, f, re.IGNORECASE):
                full = os.path.join(base, f)
                matches.append(full)
    return matches


def read_txt_files(file_list, lot, wafer):
    """
    Reads and concatenates data from a list of .txt files.
    """
    data_list = []
    for fp in file_list:
        if 'ReferenceChipSingleBeam' in fp:
            continue
        if not fp.lower().endswith('.txt'):
            continue
        try:
            df = pd.read_csv(fp, delimiter=',', header=0, skipinitialspace=True)
            if all(c in df.columns for c in ['X', 'Y', 'Z']):
                sub = df[['X', 'Y', 'Z']].copy()
                sub['Lot'] = str(lot)  # Ensure Lot is treated as a string
                sub['Wafer'] = str(wafer)  # Ensure Wafer is treated as a string
                data_list.append(sub)
                logging.debug(f"read_txt_files: appended data from {fp} => lot={lot}, wafer={wafer}")
            else:
                logging.warning(f"{fp} missing X,Y,Z.")
        except Exception as e:
            logging.error(f"Error reading {fp}: {e}")
    return pd.concat(data_list, ignore_index=True) if data_list else pd.DataFrame()


def analyze_data(data, enable_outlier_filter=False, enable_radius=False, radius=95, compute_actual_center=False, z_threshold=3.0):
    """
    Analyzes the data to compute standard deviation, bow, and warp.
    """
    if data.empty:
        return None
    pts = data[['X', 'Y']].values
    vals_mm = data['Z'].values
    coeffs = fit_plane(pts, vals_mm, enable_radius=enable_radius, radius=radius)
    corrected = remove_plane(pts, vals_mm, coeffs)
    std_dev = np.std(vals_mm)
    bow, warp = calculate_bow_warp(
        pts, corrected,
        enable_outlier_filter=enable_outlier_filter,
        compute_actual_center=compute_actual_center,
        z_threshold=z_threshold
    )
    return std_dev, bow, warp, pts, corrected


# ------------------------------------------------------------------------------
# 6) Custom Save Routines (No prompt)
# ------------------------------------------------------------------------------
def save_combined_data_autoname(big_df, output_folder, base_name):
    """
    Saves combined X,Y,Z data to <output_folder>/<base_name>.csv
    """
    if not output_folder:
        logging.warning("No output folder specified for combined data.")
        return
    if not base_name:
        base_name = "xyz_data"

    fname = os.path.join(output_folder, f"{base_name}.csv")
    if 'Wafer' in big_df.columns:
        big_df['Wafer'] = big_df['Wafer'].astype(str)  # Ensure Wafer is treated as string
    if 'Lot' in big_df.columns:
        big_df['Lot'] = big_df['Lot'].astype(str)  # Ensure Lot is treated as string
    if 'Lot' in big_df.columns and 'Wafer' in big_df.columns:
        big_df.sort_values(by=['Lot', 'Wafer'], inplace=True, ignore_index=True)

    big_df.to_csv(fname, index=False, encoding='utf-8-sig')
    logging.info(f"Combined data saved as {fname}")


def save_results_autoname(results, threshold, control_limit, output_folder, base_name):
    """
    Saves Bow/Warp results to <output_folder>/<base_name>.xlsx
    Applies conditional formatting:
        - Red for values >= USL (threshold) or <= -USL
        - Yellow for values between Control Limit and USL
        - Yellow for values between -USL and -Control Limit
    """
    if not output_folder:
        logging.warning("No output folder for results.")
        return
    if not base_name:
        base_name = "analysis_results"

    fname = os.path.join(output_folder, f"{base_name}.xlsx")
    df = pd.DataFrame(results)
    if 'Wafer' in df.columns:
        df['Wafer'] = pd.to_numeric(df['Wafer'], errors='coerce')
    if 'Lot' in df.columns:
        df['Lot'] = df['Lot'].astype(str)
    if 'Lot' in df.columns and 'Wafer' in df.columns:
        df.sort_values(by=['Lot', 'Wafer'], inplace=True, ignore_index=True)

    df.to_excel(fname, index=False)
    wb = load_workbook(fname)
    ws = wb.active

    bow_col, warp_col = None, None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Bow (µm)":
            bow_col = col[0].column_letter
        elif col[0].value == "Warp (µm)":
            warp_col = col[0].column_letter

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    if bow_col:
        # Red fill for Bow >= USL
        ws.conditional_formatting.add(
            f"{bow_col}2:{bow_col}{ws.max_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=[str(threshold)], fill=red_fill)
        )
        # Red fill for Bow <= -USL
        ws.conditional_formatting.add(
            f"{bow_col}2:{bow_col}{ws.max_row}",
            CellIsRule(operator="lessThanOrEqual", formula=[str(-threshold)], fill=red_fill)
        )
        # Yellow fill for Control Limit <= Bow < USL
        ws.conditional_formatting.add(
            f"{bow_col}2:{bow_col}{ws.max_row}",
            CellIsRule(operator="between", formula=[str(control_limit), str(threshold - 0.0001)], fill=yellow_fill)
        )
        # Yellow fill for -USL < Bow <= -Control Limit
        ws.conditional_formatting.add(
            f"{bow_col}2:{bow_col}{ws.max_row}",
            CellIsRule(operator="between", formula=[str(-threshold + 0.0001), str(-control_limit)], fill=yellow_fill)
        )
    if warp_col:
        # Red fill for Warp >= USL
        ws.conditional_formatting.add(
            f"{warp_col}2:{warp_col}{ws.max_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=[str(threshold)], fill=red_fill)
        )
        # Red fill for Warp <= -USL
        ws.conditional_formatting.add(
            f"{warp_col}2:{warp_col}{ws.max_row}",
            CellIsRule(operator="lessThanOrEqual", formula=[str(-threshold)], fill=red_fill)
        )
        # Yellow fill for Control Limit <= Warp < USL
        ws.conditional_formatting.add(
            f"{warp_col}2:{warp_col}{ws.max_row}",
            CellIsRule(operator="between", formula=[str(control_limit), str(threshold - 0.0001)], fill=yellow_fill)
        )
        # Yellow fill for -USL < Warp <= -Control Limit
        ws.conditional_formatting.add(
            f"{warp_col}2:{warp_col}{ws.max_row}",
            CellIsRule(operator="between", formula=[str(-threshold + 0.0001), str(-control_limit)], fill=yellow_fill)
        )

    wb.save(fname)
    logging.info(f"Results saved as {fname}")


# ------------------------------------------------------------------------------
# 7) GUI Class Definition
# ------------------------------------------------------------------------------
class FolderSelectorApp:
    def __init__(self, root):
        self.root = root
        self.root.geometry("700x700")
        self.root.title("Warpspeed Analysis Tool v2.04")

        self.selected_folders = []
        self.selected_files = []
        self.output_folder = None
        self.heatmap_var = IntVar(value=1)
        self.interp_var = tk.StringVar(value="cubic")

        # New variables for additional features
        self.outlier_filter_var = IntVar(value=0)
        self.radius_correction_var = IntVar(value=0)
        self.radius_value_var = tk.StringVar(value="80")  # Set default radius to 80
        self.actual_center_var = IntVar(value=0)
        self.zscore_threshold_var = tk.StringVar(value="0")  # Default Z-score threshold set to 0

        self.create_widgets()
        self.worker_thread = None

    def create_widgets(self):
        # Top frame => Add folder, add file
        top_frame = Frame(self.root)
        top_frame.pack(pady=5)

        Button(top_frame, text="Add Folder", command=self.add_folder).grid(row=0, column=0, padx=5)
        Button(top_frame, text="Add Text File", command=self.add_file).grid(row=0, column=1, padx=5)

        Label(self.root, text="Selected Folders / Files:").pack()
        f_frame = Frame(self.root)
        f_frame.pack()
        self.listbox = Listbox(f_frame, width=100, height=3)
        self.listbox.pack(side="left", fill="both", expand=True)
        sb = Scrollbar(f_frame, orient="vertical", command=self.listbox.yview)
        sb.pack(side="right", fill="y")
        self.listbox.config(yscrollcommand=sb.set)

        Button(self.root, text="Remove Selected", command=self.remove_item).pack(pady=5)

        # USL and Control Limit
        thr_frame = Frame(self.root)
        thr_frame.pack(pady=5)
        Label(thr_frame, text="USL [µm]:").pack(side="left")
        self.threshold_entry = Entry(thr_frame, width=10)
        self.threshold_entry.pack(side="left", padx=5)
        self.threshold_entry.insert(0, "100.0")

        Label(thr_frame, text="Control Limit [µm]:").pack(side="left")
        self.control_limit_entry = Entry(thr_frame, width=10)
        self.control_limit_entry.pack(side="left", padx=5)
        self.control_limit_entry.insert(0, "50.0")  # Default Control Limit

        # Output folder
        outf = Frame(self.root)
        outf.pack(pady=5)
        Label(outf, text="Output Folder:").pack(side="left")
        Button(outf, text="Select Folder", command=self.select_output_folder).pack(side="left", padx=5)

        # Checkbutton => 2D Heatmaps
        Checkbutton(self.root, text="Generate 2D Heatmaps", variable=self.heatmap_var).pack(pady=5)

        # Interpolation
        interp_frame = Frame(self.root)
        interp_frame.pack(pady=5)
        Label(interp_frame, text="2D Heatmap Interpolation:").pack(side="left")
        self.interp_combo = ttk.Combobox(interp_frame, textvariable=self.interp_var,
                                         values=["cubic", "linear"], width=10)
        self.interp_combo.pack(side="left", padx=5)
        self.interp_combo.current(0)

        # Additional Features
        features_frame = LabelFrame(self.root, text="Advanced Features", padx=10, pady=10)
        features_frame.pack(pady=10, padx=10, fill="x")

        # Outlier Filtering
        self.outlier_check = Checkbutton(features_frame, text="Enable Outlier Filtering",
                                         variable=self.outlier_filter_var,
                                         command=self.toggle_outlier_entry)  # Added command
        self.outlier_check.grid(row=0, column=0, sticky="w", pady=2, padx=(0, 5))

        # Z-Score Threshold for Outlier Filtering
        Label(features_frame, text="Z-Score Threshold:").grid(row=0, column=1, sticky="e", padx=(5, 2))
        self.zscore_entry = Entry(features_frame, textvariable=self.zscore_threshold_var, width=10, state="disabled")
        self.zscore_entry.grid(row=0, column=2, padx=(2, 5))
        self.zscore_threshold_var.set("0")  # Ensure it's set to 0 initially

        # Radius-Based Tilt Correction
        self.radius_check = Checkbutton(features_frame, text="Enable Radius-Based Tilt Correction",
                                        variable=self.radius_correction_var,
                                        command=self.toggle_radius_entry)
        self.radius_check.grid(row=1, column=0, sticky="w", pady=2)

        self.radius_entry = Entry(features_frame, textvariable=self.radius_value_var, width=10, state="disabled")
        self.radius_entry.grid(row=1, column=1, padx=5)
        Label(features_frame, text="Radius (≤ 95)").grid(row=1, column=2, sticky="w")

        # Compute Actual Center
        self.center_check = Checkbutton(features_frame, text="Enable Actual Center Computation",
                                        variable=self.actual_center_var)
        self.center_check.grid(row=2, column=0, sticky="w", pady=2)

        # Add text fields for the file base names
        file_label_frame = Frame(self.root)
        file_label_frame.pack(pady=5)

        Label(file_label_frame, text="XYZ File Name:").grid(row=0, column=0, sticky="e", padx=5)
        self.xyz_entry = Entry(file_label_frame, width=25)
        self.xyz_entry.insert(0, "xyz")  # default
        self.xyz_entry.grid(row=0, column=1, padx=5)

        Label(file_label_frame, text="Bow/Warp Analysis:").grid(row=1, column=0, sticky="e", padx=5)
        self.analysis_entry = Entry(file_label_frame, width=25)
        self.analysis_entry.insert(0, "analysis_results")  # default
        self.analysis_entry.grid(row=1, column=1, padx=5)

        # Description text
        self.desc_txt = Text(self.root, height=4, width=100, wrap="word")
        self.desc_txt.pack(pady=5)
        self.desc_txt.insert(
            "end",
            "• Bow is (mean of plane-corrected surface) - center, in microns.\n"
            "• Warp is the peak-to-peak range of the plane-corrected surface, in microns.\n"
            "• Advanced features allow for outlier filtering and dynamic center computation for more accurate analysis.\n"
        )
        self.desc_txt.config(state="disabled")

        # Progress
        self.prog_label = Label(self.root, text="Progress: 0%")
        self.prog_label.pack()
        self.prog_bar = ttk.Progressbar(self.root, orient="horizontal", length=700, mode="determinate")
        self.prog_bar.pack(pady=5)

        # Bottom
        bot_frame = Frame(self.root)
        bot_frame.pack(pady=5)
        self.run_btn = Button(bot_frame, text="Run Analysis", command=self.start_analysis, width=15)
        self.run_btn.pack(side="left", padx=10)
        Button(bot_frame, text="Exit", command=self.root.quit, width=10).pack(side="left", padx=10)

        Label(self.root, text="Reference Memo: lund-1", fg="gray").pack(side="bottom", pady=5)

    # ---------------------
    # Folder/File Add/Remove
    # ---------------------
    def add_folder(self):
        folder = filedialog.askdirectory(mustexist=True, title="Select Folder")
        if folder and folder not in self.selected_folders:
            self.selected_folders.append(folder)
            self.listbox.insert("end", "[Folder] " + folder)
            logging.info(f"Add Folder => {folder}")

    def add_file(self):
        f = filedialog.askopenfilename(title="Select .txt File", filetypes=[("Text Files", "*.txt")])
        if f and f not in self.selected_files:
            self.selected_files.append(f)
            self.listbox.insert("end", "[File] " + f)
            logging.info(f"Add File => {f}")

    def remove_item(self):
        sel = self.listbox.curselection()
        for i in reversed(sel):
            val = self.listbox.get(i)
            if val.startswith("[Folder]"):
                path = val.replace("[Folder] ", "")
                if path in self.selected_folders:
                    self.selected_folders.remove(path)
            elif val.startswith("[File]"):
                path = val.replace("[File] ", "")
                if path in self.selected_files:
                    self.selected_files.remove(path)
            self.listbox.delete(i)

    def select_output_folder(self):
        self.output_folder = filedialog.askdirectory(mustexist=True, title="Select Output Folder")
        if self.output_folder:
            logging.info(f"Selected output folder: {self.output_folder}")

    # ---------------------
    # Toggle Radius Entry
    # ---------------------
    def toggle_radius_entry(self):
        if self.radius_correction_var.get() == 1:
            self.radius_entry.config(state="normal")
        else:
            self.radius_entry.config(state="disabled")

    # ---------------------
    # Toggle Outlier Entry
    # ---------------------
    def toggle_outlier_entry(self):
        if self.outlier_filter_var.get() == 1:
            self.zscore_entry.config(state="normal")
            self.zscore_threshold_var.set("3.0")  # Set to default when enabling
        else:
            self.zscore_threshold_var.set("0")
            self.zscore_entry.config(state="disabled")

    # ---------------------
    # Progress
    # ---------------------
    def reset_progress(self):
        self.prog_label.config(text="Progress: 0%")
        self.prog_bar["value"] = 0
        self.root.update_idletasks()

    def update_progress(self, current, total):
        pct = int((current / total) * 100) if total > 0 else 0
        self.prog_label.config(text=f"Progress: {pct}%")
        self.prog_bar["value"] = pct
        self.root.update_idletasks()

    # ---------------------
    # Main Analysis
    # ---------------------
    def start_analysis(self):
        if not self.selected_folders and not self.selected_files:
            messagebox.showerror("Error", "No folders/files selected.")
            return
        if not self.output_folder:
            messagebox.showerror("Error", "No output folder selected.")
            return

        # Disable run button to prevent multiple clicks
        self.run_btn.config(state="disabled")
        self.worker_thread = threading.Thread(target=self.run_analysis)
        self.worker_thread.start()

    def run_analysis(self):
        thr_str = self.threshold_entry.get()
        try:
            threshold = float(thr_str)
        except ValueError:
            threshold = 100.0  # default
            logging.warning("Invalid threshold value entered. Using default 100.0 microns.")

        ctrl_str = self.control_limit_entry.get()
        try:
            control_limit = float(ctrl_str)
        except ValueError:
            control_limit = threshold * 0.5  # default to 0.5 of USL
            logging.warning("Invalid control limit value entered. Using default 0.5 of USL.")

        # Retrieve Z-score threshold from GUI
        zthr_str = self.zscore_threshold_var.get()
        try:
            z_threshold = float(zthr_str)
            if z_threshold <= 0 and self.outlier_filter_var.get() == 1:
                raise ValueError
        except ValueError:
            if self.outlier_filter_var.get() == 1:
                z_threshold = 3.0  # Default value
                logging.warning("Invalid Z-score threshold entered. Using default 3.0.")
            else:
                z_threshold = 0.0
                logging.info("Outlier filtering disabled. Z-score threshold set to 0.")

        # Counters for values above Control Limit and USL
        bow_above_control_limit = 0
        bow_above_usl = 0
        warp_above_control_limit = 0
        warp_above_usl = 0

        # Gather user-supplied filenames
        xyz_basename = self.xyz_entry.get().strip()
        analysis_basename = self.analysis_entry.get().strip()

        # Gather options from GUI
        enable_outlier_filter = bool(self.outlier_filter_var.get())
        enable_radius = bool(self.radius_correction_var.get())
        try:
            radius = float(self.radius_value_var.get()) if enable_radius else 95.0
            if enable_radius and radius > 95:
                logging.warning("Radius value exceeds 95. Setting radius to 95.")
                radius = 95.0
        except ValueError:
            radius = 95.0
            logging.warning("Invalid radius value entered. Using default 95.")

        compute_actual_center = bool(self.actual_center_var.get())

        logging.info(f"Analysis Options - Outlier Filtering: {enable_outlier_filter}, "
                     f"Radius-Based Correction: {enable_radius} (Radius: {radius}), "
                     f"Actual Center Computation: {compute_actual_center}, "
                     f"Z-Score Threshold: {z_threshold}")

        items = self.gather_wafer_items()
        total = len(items)
        if total == 0:
            logging.info("No wafer items found after gather_wafer_items.")
            self.finish_analysis(bow_above_control_limit, bow_above_usl, warp_above_control_limit, warp_above_usl)
            return

        all_frames = []
        results = []
        hm_info = []

        # Use ThreadPoolExecutor for data processing
        with ThreadPoolExecutor() as ex:
            future_map = {}
            for witem in items:
                fut = ex.submit(self.process_item, witem, enable_outlier_filter, enable_radius, radius,
                                compute_actual_center, z_threshold)
                future_map[fut] = witem

            done_cnt = 0
            for fut in as_completed(future_map):
                try:
                    out = fut.result()
                    if out:
                        df_data, wafer_res, hm_data = out
                        if not df_data.empty:
                            all_frames.append(df_data)
                        if wafer_res:
                            results.append(wafer_res)
                            # Update counters using absolute values
                            if abs(wafer_res["Bow (µm)"]) >= control_limit:
                                bow_above_control_limit += 1
                            if abs(wafer_res["Bow (µm)"]) >= threshold:
                                bow_above_usl += 1
                            if abs(wafer_res["Warp (µm)"]) >= control_limit:
                                warp_above_control_limit += 1
                            if abs(wafer_res["Warp (µm)"]) >= threshold:
                                warp_above_usl += 1
                        if hm_data:
                            hm_info.append(hm_data)
                except Exception as e:
                    logging.error(f"Error analyzing item: {e}")
                done_cnt += 1
                self.update_progress(done_cnt, total)

        # Save combined data => no user prompt, use "xyz_basename"
        if all_frames:
            comb_df = pd.concat(all_frames, ignore_index=True)
            save_combined_data_autoname(comb_df, self.output_folder, xyz_basename)

        # Save numeric results => pass control_limit to handle conditional formatting
        if results:
            save_results_autoname(results, threshold, control_limit, self.output_folder, analysis_basename)

        # Heatmaps
        if self.heatmap_var.get() == 1 and hm_info:
            self.reset_progress()
            total_maps = len(hm_info)
            current_map = 0
            chosen_method = self.interp_var.get()

            # Use ThreadPoolExecutor for asynchronous heatmap generation
            with ThreadPoolExecutor() as heatmap_ex:
                heatmap_futures = []
                for (lot, wafer, pts, corr_vals, bow, warp) in hm_info:
                    fut = heatmap_ex.submit(generate_heatmap, pts, corr_vals, bow, warp, lot, wafer,
                                           self.output_folder, chosen_method)
                    heatmap_futures.append(fut)

                for fut in as_completed(heatmap_futures):
                    try:
                        fut.result()
                    except Exception as e:
                        logging.error(f"Error generating heatmap: {e}")
                    current_map += 1
                    self.update_progress(current_map, total_maps)

        self.finish_analysis(bow_above_control_limit, bow_above_usl, warp_above_control_limit, warp_above_usl)

    def finish_analysis(self, bow_above_control_limit, bow_above_usl, warp_above_control_limit, warp_above_usl):
        self.update_progress(100, 100)
        logging.info("Analysis complete.")

        # Create a custom dialog with bold "Analysis Complete" and spacing
        self.show_custom_dialog(bow_above_control_limit, bow_above_usl,
                                warp_above_control_limit, warp_above_usl)

        self.run_btn.config(state="normal")
        # Reset progress bar after analysis completes
        self.reset_progress()

    def show_custom_dialog(self, bow_above_control_limit, bow_above_usl,
                           warp_above_control_limit, warp_above_usl):
        dialog = tk.Toplevel(self.root)
        dialog.title("Analysis Complete")
        dialog.geometry("400x300")
        dialog.grab_set()  # Make the dialog modal

        # Define bold font
        bold_font = tkfont.Font(dialog, weight="bold", size=12)

        # Analysis Complete Label
        analysis_label = Label(dialog, text="Analysis Complete", font=bold_font)
        analysis_label.pack(pady=(20, 10))

        # Spacer
        spacer = Label(dialog, text="")
        spacer.pack()

        # Bow and Warp Counts
        bow_control_label = Label(dialog, text=f"Bow values above Control Limit: {bow_above_control_limit}")
        bow_control_label.pack(pady=2)

        bow_usl_label = Label(dialog, text=f"Bow values above USL: {bow_above_usl}")
        bow_usl_label.pack(pady=2)

        warp_control_label = Label(dialog, text=f"Warp values above Control Limit: {warp_above_control_limit}")
        warp_control_label.pack(pady=2)

        warp_usl_label = Label(dialog, text=f"Warp values above USL: {warp_above_usl}")
        warp_usl_label.pack(pady=2)

        # Close Button
        close_btn = Button(dialog, text="Close", command=dialog.destroy)
        close_btn.pack(pady=20)

    # ---------------------
    # Gathering Wafers
    # ---------------------
    def gather_wafer_items(self):
        encountered = {}

        # Process selected folders
        for fdir in self.selected_folders:
            logging.debug(f"[gather_wafer_items] Folder => {fdir}")
            for root, dirs, _ in os.walk(fdir):
                for d in dirs:
                    subp = os.path.join(root, d)
                    logging.debug(f"Processing subfolder: {subp}")
                    lot, wafer = smart_extract_lot_wafer(subp)
                    logging.debug(f"Extracted lot={lot}, wafer={wafer} from {subp}")

                    if not lot:
                        continue
                    txt_files = sieve(subp, r'Run\d+\.txt')
                    if not txt_files:
                        continue

                    # If wafer is None, attempt to find it in sub-subfolders
                    if wafer is None:
                        deeper_wafer = self.check_subfolders_for_wafer(subp)
                        if deeper_wafer:
                            logging.debug(
                                f"Skipping partial lot={lot} subp={subp} because deeper subfolder has wafer={deeper_wafer}")
                            continue
                        wafer = "Unknown"

                    # Ensure lot and wafer are strings
                    key = (str(lot), str(wafer))

                    if key not in encountered:
                        encountered[key] = set()
                    for fp in txt_files:
                        encountered[key].add(fp)

        # Process selected files
        for single_fp in self.selected_files:
            logging.debug(f"[gather_wafer_items] Single file => {single_fp}")
            start_dir = os.path.dirname(single_fp)
            lot, wafer = find_lot_wafer_upwards(start_dir)
            logging.debug(f"Found lot={lot}, wafer={wafer} for file {single_fp}")
            if not lot:
                lot = "Unknown"
            if not wafer:
                wafer = "Unknown"
            key = (str(lot), str(wafer))

            if key not in encountered:
                encountered[key] = set()
            encountered[key].add(single_fp)

        witems = []
        for (lot, wafer), filepaths in encountered.items():
            logging.debug(f"Final dictionary => lot={lot}, wafer={wafer}, #files={len(filepaths)}")
            witems.append({
                'lot': lot,
                'wafer': wafer,
                'files': list(filepaths)
            })
        return witems

    # ---------------------
    # Defining check_subfolders_for_wafer as a method inside the class
    # ---------------------
    def check_subfolders_for_wafer(self, folder):
        run_folder_pattern = re.compile(r'^Run\d+$', re.IGNORECASE)  # Pattern to match 'Run' folders
        for root, dirs, _ in os.walk(folder):
            for d in dirs:
                if run_folder_pattern.match(d):
                    wafer = os.path.basename(root)  # Parent folder is the wafer
                    logging.debug(f"Found wafer={wafer} before Run folder {os.path.join(root, d)}")
                    return wafer
        logging.debug(f"No Run folders found in folder: {folder}")
        return None

    def process_item(self, wdict, enable_outlier_filter, enable_radius, radius, compute_actual_center, z_threshold):
        """
        Processes a single wafer item: reads data, analyzes it, and prepares heatmap information.
        """
        lot = wdict['lot']
        wafer = wdict['wafer']
        flist = wdict['files']
        logging.debug(f"[process_item] lot={lot}, wafer={wafer}, #files={len(flist)}")

        df_data = read_txt_files(flist, str(lot), str(wafer))  # Ensure passing as strings
        if df_data.empty:
            logging.info(f"No valid data for {lot}/{wafer}")
            return None

        result = analyze_data(
            df_data,
            enable_outlier_filter=enable_outlier_filter,
            enable_radius=enable_radius,
            radius=radius,
            compute_actual_center=compute_actual_center,
            z_threshold=z_threshold
        )
        if result:
            std_dev, bow, warp, pts, corr = result
            wafer_res = {
                "Lot": lot,
                "Wafer": wafer,
                "Std Dev": std_dev,
                "Bow (µm)": bow,
                "Warp (µm)": warp
            }
            hm = (lot, wafer, pts, corr, bow, warp)
            return (df_data, wafer_res, hm)
        return None


# ------------------------------------------------------------------------------
# 8) Running the GUI
# ------------------------------------------------------------------------------
def run_gui():
    root = tk.Tk()
    app = FolderSelectorApp(root)
    root.mainloop()


if __name__ == "__main__":
    run_gui()
